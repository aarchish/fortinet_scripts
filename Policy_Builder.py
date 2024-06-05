import openpyxl as xl;

filePath = input("Enter Excel file path: ");
print("Following column values are mandatory and cells cannot be empty");
print("Column A : Policy Name ");
print("Column B : Destination Address");
print("Column C : Comments for Firewall Object");
print("Column D : Service Ports for Firewall Policy");

wb_obj = xl.load_workbook(filePath);
sheet_obj = wb_obj.active;

op = input("Enter output file name with .txt (eg: output.txt) (new file will be created): ");
file = open(op, "x");
file.writelines("config firewall policy\n");


service_dict = {};
custom_services = open("Services_Script.txt", "x");
custom_services.writelines("config firewall service custom\n");

#for col in sheet_obj.iter_cols(1, sheet_obj.max_column):
for row in sheet_obj.iter_rows(2, sheet_obj.max_row):
    file.writelines("edit 0\n");
    file.writelines("set name \""+row[0].value+"\"\n");
    file.writelines("set srcintf \"extvxlan\" \"intvxlan\"\n");
    file.writelines("set dstintf \"extvxlan\" \"intvxlan\"\n");
    file.writelines("set action accept\n");
    file.writelines("set srcaddr CLOUDFLARE_IP\n");
    file.writelines("set dstaddr "+row[1].value+"\n");
    file.writelines("set schedule always\n");
    file.writelines("set utm-status enable\n");
    file.writelines("set profile-type group\n");
    file.writelines("set profile-group CSX_INGRESS_Profile_Grp\n");
    file.writelines("set logtraffic all\n");
    file.writelines("set comments \""+row[2].value+"\"\n");
    service_line = "set service ";
    services = str(row[3].value);
    services = services.split("/");
    for service in services:
        if(service == "21"):
            service_line += " "+"FTP";
        elif(service == "22"):
            service_line += " "+"SSH";
        elif(service == "25"):
            service_line += " "+"SMTP";
        elif(service == "53"):
            service_line += " "+"DNS";
        elif(service == "80"):
            service_line += " "+"HTTP";
        elif(service == "143"):
            service_line += " "+"IMAP";
        elif(service == "389"):
            service_line += " "+"LDAP";
        elif(service == "443"):
            service_line += " "+"HTTPS";
        elif(service == "465"):
            service_line += " "+"SMTPS";
        elif(service == "993"):
            service_line += " "+"IMAPS";
        else:
            new_service = "\""+service+"-TCP\"";
            if service_dict.get(service):
                service_line += " "+new_service;
            else:
                service_line += " "+new_service;
                service_dict[service] = "1";
                custom_services.writelines("edit "+new_service+"\n");
                custom_services.writelines("set tcp-portrange "+service+"\n");
                custom_services.writelines("next\n");
    
    file.writelines(service_line+"\n");
    file.writelines("next\n");

file.writelines("end\n");
file.close();

custom_services.writelines("end\n");
custom_services.close();



