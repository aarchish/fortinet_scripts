
import pandas as pd;
import openpyxl as xl;


filePath = input("Enter Excel file path: ");
print("Following column values are mandatory and cells cannot be empty");
print("Column A : Object Name Variable ");
print("Column B : Type of Address Variable - Subnet or FQDN");
print("Column C : Address Object Variable");
print("Column D : Comments for Firewall Object");

wb_obj = xl.load_workbook(filePath);
sheet_obj = wb_obj.active;

name_prefix_required = input("is a prefix required for all the objects to be created?(True/False) : ");
name_prefix = "";
if(name_prefix_required == "True") : name_prefix = input("Enter the common name prefix that is required : ")

op = input("Enter output file name with .txt (eg: output.txt) (new file will be created): ");
file = open(op, "x");
file.writelines("config firewall address\n");

#for col in sheet_obj.iter_cols(1, sheet_obj.max_column):
for row in sheet_obj.iter_rows(2, sheet_obj.max_row):
    object_name = ("edit "+name_prefix+row[0].value) if (name_prefix_required) else ("edit "+row[0].value);
    object_type = "set type " +row[1].value;
    object_address = row[1].value+" "+row[2].value;
    object_comment = "set comment "+row[3].value;

    file.writelines(object_name+'\n');
    #file.writelines('\n');
    file.writelines(object_type+'\n');
    #file.writelines('\n'+'\n');
    file.writelines(object_address+'\n');
    #file.writelines('\n');
    file.writelines(object_comment+'\n');
    #file.writelines('\n'); 
    file.writelines("next\n");
file.writelines("end\n");
file.close();



