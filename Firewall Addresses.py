
import pandas as pd;
import openpyxl as xl;

#df = pd.read_excel("Subnets.xlsx") 
#print(df)

filePath = input("Enter Excel file path: ");
print("Following column values are mandatory and cells cannot be empty");
print("Column A : Object Name Variable ");
print("Column B : Type of Address Variable - Subnet or FQDN");
print("Column C : Address Object Variable");
print("Column D : Comments for Firewall Object");

#filePath = "Subnets.xlsx";
wb_obj = xl.load_workbook(filePath);
sheet_obj = wb_obj.active;
#print(sheet_obj.max_row);
#print(sheet_obj.max_column);
name_prefix_required = input("is a prefix required for all the objects to be created?(True/False) : ");
name_prefix = "";
if(name_prefix_required == "True") : name_prefix = input("Enter the common name prefix that is required : ")

op = input("Enter output file name with .txt (eg: output.txt) (new file will be created): ");
#op = "output1.txt";
file = open(op, "x");
#L = ["This is Lagos \n","This is Python \n","This is Fcc \n"]
#file.writelines(L);
file.writelines("config firewall address\n");

#for col in sheet_obj.iter_cols(1, sheet_obj.max_column):
for row in sheet_obj.iter_rows(2, sheet_obj.max_row):
    if(name_prefix_required) : 
        object_name = "edit "+name_prefix+row[0].value;     
    else: object_name = "edit "+row[0].value;
    object_type = "set type " +row[1].value;
    object_address = row[1].value+" "+row[2].value;
    object_comment = "set comment "+row[3].value;

    file.writelines(object_name+'\n');
    #file.writelines('\n');
    file.writelines(object_type+'\n');
    #file.writelines('\n'+'\n');
    file.writelines(object_address+'\n');
    #file.writelines('\n');
    file.writelines(object_comment+'\n');
    #file.writelines('\n'); 
    file.writelines("next\n");
file.writelines("end\n");
file.close();



